#!/usr/bin/env python3
"""
build_client_sheet.py — Build the client-facing "Todays Performance" Excel
from a cleaned + tagged Sprinklr daily export.

Mirrors the layout of the real client deliverable exactly:
  - Banner rows 2-9 (date headers + benchmark descriptors)
  - Header at row 11: Brand | Title | Publish Date | Channel | Post Type |
    Text | Engagements | ER | Benchmark Δ | Pacing
  - Data from row 12, organic only (paid excluded), sorted by Benchmark Δ desc
  - Hyperlinks on the Channel cell (matches real file convention)
  - Publish dates stripped to date-only (no time)
  - ER displayed with '0%' number format
  - Multi-level benchmark fallback when J9:M21 doesn't have a (Brand, Channel)

Usage:
    python build_client_sheet.py \\
        --export /path/to/DailyBoostablePostsAutomatedExport.xlsx \\
        --memory-bank /path/to/memory_bank.csv \\
        --reporting-date 2026-05-26 \\
        --output ./LF_Amazon_Boostable_Posts.xlsx \\
        [--benchmarks /path/to/benchmarks.csv]   # optional, falls back to internal defaults

The benchmarks CSV (optional) has three columns: brand,channel,benchmark_engagements.
If omitted, the script uses the built-in BENCHMARK table (J9:M21 snapshot from
the V2 internal sheet). For best accuracy, export the J9:M21 table fresh after
each Friday refresh and pass it here.
"""

from __future__ import annotations
import argparse, csv, sys
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Optional

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("openpyxl is required. Install with: pip install openpyxl")


# Default benchmarks (J9:M21 snapshot). Override via --benchmarks CSV.
DEFAULT_BENCHMARKS = {
    ('Prime Video', 'Facebook'): 18206.91,
    ('Prime Video', 'Instagram'): 54918.87,
    ('Prime Video', 'TikTok'): 61501.54,
    ('Prime Video', 'YouTube'): 8489.04,
    ('Culture Rated', 'Facebook'): 1355.63,
    ('Culture Rated', 'Instagram'): 43936.14,
    ('Culture Rated', 'TikTok'): 118280.17,
    ('Prime Movies', 'Facebook'): 14365.84,
    ('Prime Movies', 'Instagram'): 16848.56,
    ('Prime Movies', 'TikTok'): 25610.26,
    ('Primero Latino', 'Instagram'): 20477.74,
    ('Primero Latino', 'TikTok'): 21288.80,
}

BRAND_MAP = {'PV United States': 'Prime Video'}
CHANNEL_MAP = {
    'TT': 'TikTok', 'IG': 'Instagram', 'FB': 'Facebook', 'YT': 'YouTube',
    'TIKTOK_BUSINESS': 'TikTok', 'INSTAGRAM': 'Instagram',
    'FBPAGE': 'Facebook', 'YOUTUBE': 'YouTube',
}
WEEKDAY = ['Monday','Tuesday','Wednesday','Thursday','Friday','Saturday','Sunday']


def parse_args():
    p = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument('--export', required=True)
    p.add_argument('--memory-bank', required=True)
    p.add_argument('--reporting-date', required=True, help='YYYY-MM-DD')
    p.add_argument('--output', required=True)
    p.add_argument('--benchmarks', help='Optional CSV (brand,channel,benchmark_engagements). '
                                        'If omitted, looks for state/benchmarks.csv next to the export.')
    p.add_argument('--state-dir', help='Directory holding benchmarks.csv (default: ../state/ relative to export)')
    p.add_argument('--auto-refresh-benchmarks', action='store_true',
                   help='Update state/benchmarks.csv from this export before building the sheet '
                        '(daily nudge to the rolling 90-day window).')
    return p.parse_args()


def load_benchmarks(path: Optional[str]) -> tuple[dict, str]:
    """Returns (benchmarks_dict, source_label) so we can tell the user where they came from."""
    if not path:
        return dict(DEFAULT_BENCHMARKS), 'built-in defaults'
    if not Path(path).exists():
        return dict(DEFAULT_BENCHMARKS), f'built-in defaults (no file at {path})'
    out, refreshed = {}, ''
    with open(path, newline='', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        for r in reader:
            out[(r['brand'].strip(), r['channel'].strip())] = float(r['benchmark_engagements'])
            refreshed = r.get('refreshed_at', refreshed)
    label = f'{path}' + (f' (refreshed {refreshed})' if refreshed else '')
    return out, label


def load_memory_bank(path: str) -> dict:
    out = {}
    p = Path(path)
    if p.suffix.lower() == '.csv':
        with p.open(newline='', encoding='utf-8-sig') as f:
            for i, r in enumerate(csv.reader(f)):
                if i == 0: continue
                if len(r) >= 3 and r[1] and r[2]:
                    out[str(r[1]).strip()] = str(r[2]).strip()
    else:
        wb = openpyxl.load_workbook(p, data_only=True, read_only=True)
        ws = wb[wb.sheetnames[0]]
        for i, r in enumerate(ws.iter_rows(values_only=True)):
            if i < 2: continue
            if r and len(r) >= 3 and r[1] and r[2]:
                out[str(r[1]).strip()] = str(r[2]).strip()
    return out


def load_export(path: str) -> tuple[list[dict], list[str]]:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows_iter = ws.iter_rows(values_only=True)
    header = []
    for r in rows_iter:
        cells = [str(c).strip() if c is not None else '' for c in r]
        if any('permalink' in c.lower() for c in cells):
            header = cells
            break
    out = []
    for r in rows_iter:
        if all(v is None or v == '' for v in r): continue
        out.append({header[i] if i < len(header) else f'c{i}': v for i, v in enumerate(r)})
    return out, header


def dedupe_clean(rows: list[dict], header: list[str]) -> list[dict]:
    """Drop zero-engagement rows, dedupe by Permalink (multi-tagged → Multi-Title)."""
    nonzero = [r for r in rows if (r.get('PV_UCM_Total Engagements (SUM)') or 0) and float(r['PV_UCM_Total Engagements (SUM)']) > 0]
    groups = defaultdict(list)
    for r in nonzero: groups[r['Permalink']].append(r)
    cleaned = []
    for url, grp in groups.items():
        if len(grp) == 1:
            cleaned.append(grp[0])
        else:
            keep = dict(grp[0]); keep['PV_Titles (Outbound Message)'] = 'Multi-Title'
            cleaned.append(keep)
    return cleaned


def split_account(acct: str) -> tuple[str, str]:
    if not acct: return ('', '')
    if ' - ' in acct:
        b, c = acct.rsplit(' - ', 1)
        return b.strip(), c.strip()
    return acct.strip(), ''


def post_type_from(url: str) -> str:
    if not url: return 'Other'
    u = url.lower()
    if any(x in u for x in ('reel','tiktok.com','shorts','video')): return 'Video'
    return 'Still'


def pacing(delta: Optional[float]) -> str:
    if delta is None: return ''
    if delta > 0: return 'Overperforming'
    if delta >= -0.5: return 'Has Potential'
    return 'Underperforming'


def enrich(cleaned: list[dict], memory_bank: dict, benchmarks: dict) -> list[dict]:
    enriched = []
    for r in cleaned:
        url = (r.get('Permalink') or '').strip()
        if not url: continue
        rb, rc = split_account(r.get('Account', ''))
        brand = BRAND_MAP.get(rb, rb)
        channel = CHANNEL_MAP.get(rc, rc)
        eng = float(r.get('PV_UCM_Total Engagements (SUM)') or 0)
        paid = float(r.get('PV_UCM_Total Paid Engagements (SUM)') or 0)
        impr = float(r.get('PV_UCM_Total Impressions (SUM)') or 0)
        er = eng / impr if impr else 0
        pub = r.get('PublishedTime') or r.get('ScheduledTime')
        if isinstance(pub, str):
            try: pub = datetime.fromisoformat(pub)
            except Exception: pass
        pub_date = pub.date() if isinstance(pub, datetime) else pub
        # Tag: Sprinklr PV_Titles → memory bank → blank
        tag = (r.get('PV_Titles (Outbound Message)') or '').strip() or memory_bank.get(url, '')
        bm = benchmarks.get((brand, channel))
        delta = ((eng - bm) / bm) if bm else None
        text = (r.get('Outbound Post ( Unified Analytics )') or '')
        text_clean = text.split('\n')[0].split('🎥')[0].split('📺')[0].strip()[:35]
        enriched.append({
            'Brand': brand, 'Title': tag, 'Publish Date': pub_date,
            'Channel': channel, 'Post Type': post_type_from(url),
            'Text': text_clean, 'Engagements': eng, 'ER': round(er, 4),
            'Benchmark Δ': round(delta, 2) if delta is not None else None,
            'Pacing': pacing(delta), 'Post Link': url, 'Has Paid': paid > 0,
        })
    return enriched


def fill_missing_benchmarks(organic: list[dict]) -> None:
    """Multi-level fallback: same brand → dataset mean."""
    none_p = [e for e in organic if not e['Pacing']]
    if not none_p: return
    brand_pool = defaultdict(list)
    for e in organic:
        if e['Pacing']: brand_pool[e['Brand']].append(e['Engagements'])
    all_means = [e['Engagements'] for e in organic if e['Pacing']]
    overall = (sum(all_means) / len(all_means)) if all_means else 0
    for e in none_p:
        pool = brand_pool.get(e['Brand'], [])
        bm = (sum(pool) / len(pool)) if pool else overall
        if bm:
            d = (e['Engagements'] - bm) / bm
            e['Benchmark Δ'] = round(d, 2)
            e['Pacing'] = pacing(d)


def is_unresolved(e: dict) -> bool:
    """A row is unresolved if its Title is empty, N/A, or a (needs-review) marker."""
    t = (e.get('Title') or '').strip()
    return (not t) or t.upper() == 'N/A' or t.startswith('(needs-review') or t.startswith('NEW:')


# Exact colors and fonts from the real 5/25 client file
FONT_FAMILY = 'Karla'   # Real file uses Google Karla; falls back gracefully if not installed
PACING_FILL = {
    'Overperforming':  PatternFill('solid', fgColor='B7E1CD'),   # green
    'Has Potential':   PatternFill('solid', fgColor='FFF2CC'),   # yellow
    'Underperforming': PatternFill('solid', fgColor='F4CCCC'),   # pink
}
LINK_BLUE = '1155CC'
REVIEW_YELLOW = PatternFill('solid', fgColor='FFEB9C')

# Column alignment per the real file
CENTER_COLS = {4, 5, 6, 8, 9, 10, 11}   # Publish Date, Channel, Post Type, Engagements, ER, Δ, Pacing
LEFT_COLS = {2, 3, 7}                    # Brand, Title, Text


def build_workbook(organic: list[dict], reporting_date: date, mark_unresolved: bool = False) -> openpyxl.Workbook:
    """Build the deliverable matching the real 5/25 client file styling exactly:
    Karla font, no header band, no borders, no zebra, exact Pacing pastels,
    Channel hyperlink in #1155CC, vertical=bottom, centered numeric columns."""
    from datetime import timedelta

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Todays Performance'

    KARLA = FONT_FAMILY  # 'Karla'
    bold_karla = Font(name=KARLA, bold=True)
    karla = Font(name=KARLA)
    bold_karla_12 = Font(name=KARLA, bold=True, size=12)
    link_karla = Font(name=KARLA, color=LINK_BLUE, underline='single')

    bot = Alignment(vertical='bottom')
    bot_center = Alignment(vertical='bottom', horizontal='center')
    bot_right = Alignment(vertical='bottom', horizontal='right')

    # Banner — B2 contains the full title "Boostable Posts | Daily Digest"
    ws['B2'] = 'Boostable Posts | Daily Digest'
    ws['B2'].font = bold_karla_12; ws['B2'].alignment = bot

    ws['C3'] = 'Today:'; ws['C3'].font = bold_karla; ws['C3'].alignment = bot_right
    ws['D3'] = WEEKDAY[reporting_date.weekday()]; ws['D3'].font = karla
    ws['E3'] = reporting_date; ws['E3'].font = karla

    ws['C4'] = 'Posts Published Between:'; ws['C4'].font = bold_karla; ws['C4'].alignment = bot_right
    ws['D4'] = reporting_date - timedelta(days=7); ws['D4'].font = karla
    ws['E4'] = '-'; ws['E4'].font = karla
    ws['F4'] = reporting_date - timedelta(days=1); ws['F4'].font = karla

    ws['B6'] = 'Benchmark:'; ws['C6'] = 'Average Engagements per Post, Rolling Past 90 Days'
    ws['B7'] = 'Benchmarking By:'; ws['C7'] = 'Individual Handle, by Channel (Culture Rated, Primero Latino, Prime Movies, Prime Video)'
    ws['B8'] = 'Channels:'; ws['C8'] = 'Posts on FB, IG, YT, TT'
    ws['B9'] = 'Handles:'; ws['C9'] = 'Prime Video, Culture Rated, Prime Movies, Primero Latino'
    for cell in ('B6','B7','B8','B9'):
        ws[cell].font = bold_karla; ws[cell].alignment = bot
    for cell in ('C6','C7','C8','C9'):
        ws[cell].font = karla; ws[cell].alignment = bot

    # Header row 11 — bold Karla, no fill, no border, alignment per column
    cols = ['Brand','Title','Publish Date','Channel','Post Type','Text','Engagements','ER','Benchmark Δ','Pacing']
    for j, h in enumerate(cols, start=2):
        c = ws.cell(row=11, column=j, value=h)
        c.font = bold_karla
        c.alignment = bot_center if j in CENTER_COLS else bot

    # Data — no borders, no zebra; Karla everywhere; Pacing fills only
    for i, e in enumerate(organic, start=12):
        title_for_row = e['Title'] or ''
        if mark_unresolved and is_unresolved(e):
            title_for_row = title_for_row or '(needs-review)'
        elif not mark_unresolved and (not title_for_row or title_for_row.startswith('(needs-review')):
            title_for_row = 'N/A'

        vals = [e['Brand'], title_for_row, e['Publish Date'], e['Channel'], e['Post Type'],
                e['Text'], float(e['Engagements']), e['ER'], e['Benchmark Δ'], e['Pacing']]
        for j, v in enumerate(vals, start=2):
            c = ws.cell(row=i, column=j, value=v)
            c.font = karla
            c.alignment = bot_center if j in CENTER_COLS else bot

        # Channel cell: hyperlink with #1155CC underlined Karla
        cc = ws.cell(row=i, column=5)
        cc.hyperlink = e['Post Link']
        cc.font = link_karla

        # Pacing fill — exact colors from the real file
        pc = ws.cell(row=i, column=11)
        fill = PACING_FILL.get(e['Pacing'])
        if fill: pc.fill = fill

        # Number formats
        ws.cell(row=i, column=4).number_format = 'yyyy-mm-dd'
        ws.cell(row=i, column=8).number_format = '#,##0'
        ws.cell(row=i, column=9).number_format = '0%'
        ws.cell(row=i, column=10).number_format = '0.00'

        # Yellow highlight on unresolved (review file only)
        if mark_unresolved and is_unresolved(e):
            ws.cell(row=i, column=3).fill = REVIEW_YELLOW  # Title
            ws.cell(row=i, column=2).fill = REVIEW_YELLOW  # Brand

    # Column widths matching real file (others use default)
    ws.column_dimensions['A'].width = 3.25
    ws.column_dimensions['B'].width = 17.38
    ws.column_dimensions['C'].width = 23.38
    ws.column_dimensions['I'].width = 12.63
    ws.column_dimensions['K'].width = 20.63
    ws.column_dimensions['L'].width = 3.25

    return wb


def resolve_benchmarks_path(args) -> Optional[str]:
    """Where benchmarks.csv lives. Priority: --benchmarks, --state-dir,
    then state/ near the export, else None (use defaults)."""
    if args.benchmarks: return args.benchmarks
    if args.state_dir:
        p = Path(args.state_dir) / 'benchmarks.csv'
        return str(p) if p.exists() else None
    export_parent = Path(args.export).resolve().parent
    for candidate in [export_parent / 'state', export_parent.parent / 'state']:
        if (candidate / 'benchmarks.csv').exists():
            return str(candidate / 'benchmarks.csv')
    return None


def main():
    args = parse_args()
    rd = datetime.strptime(args.reporting_date, '%Y-%m-%d').date()

    if args.auto_refresh_benchmarks:
        from subprocess import run
        state_dir = args.state_dir or str(Path(args.export).resolve().parent.parent / 'state')
        print(f'Auto-refreshing benchmarks state in {state_dir}...')
        run(['python3', str(Path(__file__).parent / 'update_benchmark_state.py'),
             '--export', args.export, '--state-dir', state_dir], check=False)

    raw, header = load_export(args.export)
    cleaned = dedupe_clean(raw, header)
    mb = load_memory_bank(args.memory_bank)
    bm_path = resolve_benchmarks_path(args)
    benchmarks, source = load_benchmarks(bm_path)
    print(f'Benchmarks loaded from: {source}')
    enriched = enrich(cleaned, mb, benchmarks)
    organic = [e for e in enriched if not e['Has Paid']]
    fill_missing_benchmarks(organic)
    organic.sort(key=lambda x: -(x['Benchmark Δ'] if x['Benchmark Δ'] is not None else -999))

    wb_client = build_workbook(organic, rd, mark_unresolved=False)
    wb_client.save(args.output)
    print(f'Wrote CLIENT file: {args.output}  ({len(organic)} organic rows, '
          f'dropped {len(enriched) - len(organic)} paid)')

    unresolved = [e for e in organic if is_unresolved(e)]
    review_path = Path(args.output).with_name(Path(args.output).stem + '_review.xlsx')
    wb_review = build_workbook(organic, rd, mark_unresolved=True)
    wb_review.save(review_path)
    print(f'Wrote REVIEW file: {review_path}  ({len(unresolved)} rows highlighted)')

    if unresolved:
        worklist = Path(args.output).with_name(f'needs-review-{rd.isoformat()}.csv')
        with worklist.open('w', newline='', encoding='utf-8') as f:
            w = csv.writer(f)
            w.writerow(['Post Link','Brand','Channel','Caption Preview','Suggested Tag'])
            for e in unresolved:
                w.writerow([e['Post Link'], e['Brand'], e['Channel'], e['Text'], e.get('Title') or '(empty)'])
        print(f'Wrote WORKLIST:    {worklist}')
        print()
        print(f'⚠  {len(unresolved)} rows need review before sending to client.')
        print('   Open the _review.xlsx, tag the yellow rows, add to memory bank, re-run.')
    else:
        print('✓  All rows tagged. No review needed.')

    from collections import Counter
    print()
    for k, v in Counter(e['Pacing'] for e in organic).most_common():
        print(f'  {k}: {v}')


if __name__ == '__main__':
    main()
