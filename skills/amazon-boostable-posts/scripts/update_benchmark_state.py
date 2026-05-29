#!/usr/bin/env python3
"""
update_benchmark_state.py — maintain a running 90-day post log + recompute
benchmarks per (Brand x Channel). Called by the daily workflow and by the
Friday 90-day refresh.

State files (written to the workspace under state/):
  - post_log.csv     append-only log of (post_date, brand, channel, permalink, engagements)
                     deduped by permalink (latest engagement count wins because
                     engagements grow over time on the same post)
  - benchmarks.csv   current benchmark per (brand, channel) computed as the
                     90-day rolling mean of organic engagements
  - benchmarks_history/benchmarks_YYYY-MM-DD.csv    daily snapshot for audit

Usage:
  # Daily run — append the 7-day daily export, recompute, write benchmarks.csv
  python update_benchmark_state.py \\
      --export /path/to/Daily Post Export - 2026-05-26.xlsx \\
      --state-dir state/

  # Friday refresh — replace the log with the 90-day export
  python update_benchmark_state.py \\
      --90day-export /path/to/90 Day Benchmark.xlsx \\
      --state-dir state/ \\
      --replace-log
"""

from __future__ import annotations
import argparse, csv, sys, shutil
from collections import defaultdict
from datetime import datetime, timedelta, date
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl required. Install with: pip install openpyxl")


BRAND_MAP = {'PV United States': 'Prime Video'}
CHANNEL_MAP = {
    'TT': 'TikTok', 'IG': 'Instagram', 'FB': 'Facebook', 'YT': 'YouTube',
    'TIKTOK_BUSINESS': 'TikTok', 'INSTAGRAM': 'Instagram',
    'FBPAGE': 'Facebook', 'YOUTUBE': 'YouTube',
}


def parse_args():
    p = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument('--export', help='Sprinklr daily export (.xlsx)')
    p.add_argument('--90day-export', dest='ninetyday', help='Sprinklr 90-day export (.xlsx)')
    p.add_argument('--state-dir', required=True, help='Where state files live (e.g. state/)')
    p.add_argument('--replace-log', action='store_true',
                   help='Wipe the existing log and seed entirely from this export (use on Friday)')
    p.add_argument('--window-days', type=int, default=90)
    return p.parse_args()


def split_account(acct: str) -> tuple[str, str]:
    if not acct: return ('', '')
    if ' - ' in acct:
        b, c = acct.rsplit(' - ', 1)
        return b.strip(), c.strip()
    return acct.strip(), ''


def load_export(path: Path) -> list[dict]:
    """Returns a list of {date, brand, channel, permalink, engagements, paid}."""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows_iter = ws.iter_rows(values_only=True)
    header = []
    for r in rows_iter:
        cells = [str(c).strip() if c is not None else '' for c in r]
        if any('permalink' in c.lower() for c in cells):
            header = cells
            break
    if not header:
        raise ValueError(f"No header in {path}")
    col = {n: i for i, n in enumerate(header)}
    out = []
    for r in rows_iter:
        if not r or all(v is None for v in r): continue
        perm = r[col['Permalink']] if col.get('Permalink') is not None else None
        if not perm: continue
        eng = float(r[col['PV_UCM_Total Engagements (SUM)']] or 0) if col.get('PV_UCM_Total Engagements (SUM)') is not None else 0
        paid_col = col.get('PV_UCM_Total Paid Engagements (SUM)')
        paid = float(r[paid_col] or 0) if paid_col is not None else 0
        if eng <= 0: continue  # drop zero-engagement rows
        if paid > 0: continue  # organic only for benchmark purposes
        pub = r[col['PublishedTime']] if col.get('PublishedTime') is not None else None
        if isinstance(pub, str):
            try: pub = datetime.fromisoformat(pub)
            except Exception: pass
        d = pub.date() if isinstance(pub, datetime) else None
        if d is None: continue
        rb, rc = split_account(r[col['Account']] if col.get('Account') is not None else '')
        brand = BRAND_MAP.get(rb, rb); channel = CHANNEL_MAP.get(rc, rc)
        if not brand or not channel: continue
        out.append({'date': d, 'brand': brand, 'channel': channel,
                    'permalink': str(perm).strip(), 'engagements': eng})
    return out


def load_log(p: Path) -> dict:
    """Returns {permalink: {date, brand, channel, engagements}}."""
    out = {}
    if not p.exists(): return out
    with p.open(newline='', encoding='utf-8') as f:
        for row in csv.DictReader(f):
            out[row['permalink']] = {
                'date': date.fromisoformat(row['date']),
                'brand': row['brand'], 'channel': row['channel'],
                'engagements': float(row['engagements']),
            }
    return out


def write_log(p: Path, log: dict):
    with p.open('w', newline='', encoding='utf-8') as f:
        w = csv.writer(f)
        w.writerow(['date','brand','channel','permalink','engagements'])
        for perm, v in log.items():
            w.writerow([v['date'].isoformat(), v['brand'], v['channel'], perm, v['engagements']])


def compute_benchmarks(log: dict, window_days: int) -> dict:
    """90-day rolling mean of organic engagements per (brand, channel)."""
    if not log: return {}
    cutoff = max(v['date'] for v in log.values()) - timedelta(days=window_days)
    buckets = defaultdict(list)
    for v in log.values():
        if v['date'] >= cutoff:
            buckets[(v['brand'], v['channel'])].append(v['engagements'])
    return {k: round(sum(vs)/len(vs), 2) for k, vs in buckets.items() if vs}


def write_benchmarks(p: Path, benchmarks: dict, refreshed: date, history_dir: Path):
    # Snapshot existing file to history (if any) before overwriting
    if p.exists():
        history_dir.mkdir(parents=True, exist_ok=True)
        snap = history_dir / f'benchmarks_{date.today().isoformat()}.csv'
        if not snap.exists():
            shutil.copy(p, snap)
    with p.open('w', newline='', encoding='utf-8') as f:
        w = csv.writer(f)
        w.writerow(['brand','channel','benchmark_engagements','refreshed_at'])
        for (brand, channel), bm in sorted(benchmarks.items()):
            w.writerow([brand, channel, bm, refreshed.isoformat()])


def main():
    args = parse_args()
    if not args.export and not args.ninetyday:
        sys.exit("Need --export or --90day-export")

    state_dir = Path(args.state_dir); state_dir.mkdir(parents=True, exist_ok=True)
    log_path = state_dir / 'post_log.csv'
    bm_path = state_dir / 'benchmarks.csv'
    history_dir = state_dir / 'benchmarks_history'

    # Load existing log (unless we're replacing)
    log = {} if args.replace_log else load_log(log_path)

    # Read new posts
    new_posts = load_export(Path(args.ninetyday or args.export))
    print(f'New posts to ingest: {len(new_posts)}', file=sys.stderr)

    # Merge: for each permalink, keep the entry with the LATEST date
    # (because engagements grow over time on the same post)
    added = updated = 0
    for p in new_posts:
        existing = log.get(p['permalink'])
        if not existing:
            log[p['permalink']] = p; added += 1
        elif p['date'] >= existing['date']:
            log[p['permalink']] = p; updated += 1

    print(f'  Added: {added}, Updated: {updated}, Total in log: {len(log)}', file=sys.stderr)

    # Trim log to window (don't keep stale data forever)
    if log:
        cutoff = max(v['date'] for v in log.values()) - timedelta(days=args.window_days + 7)
        before = len(log)
        log = {k: v for k, v in log.items() if v['date'] >= cutoff}
        if len(log) < before:
            print(f'  Trimmed {before - len(log)} entries older than {cutoff}', file=sys.stderr)

    # Persist log + recompute benchmarks
    write_log(log_path, log)
    benchmarks = compute_benchmarks(log, args.window_days)
    write_benchmarks(bm_path, benchmarks, date.today(), history_dir)

    print(f'\nBenchmarks written to {bm_path}', file=sys.stderr)
    print(f'  Refreshed: {date.today().isoformat()}', file=sys.stderr)
    print(f'  Brand×Channel combos: {len(benchmarks)}', file=sys.stderr)
    for k, v in sorted(benchmarks.items()):
        print(f'    {k[0]:<14} {k[1]:<10} {v:>12,.2f}', file=sys.stderr)


if __name__ == '__main__':
    main()
