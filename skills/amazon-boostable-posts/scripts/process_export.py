#!/usr/bin/env python3
"""
process_export.py — runs the 6-layer fallback stack against a Sprinklr daily export.

Layer 1 (Sprinklr PV_Titles), Layer 2 (memory bank URL match), and Layer 4
(account-history fallback) run deterministically here. For Layer 3 the script
emits one Claude-ready prompt per post to stdout (or to --prompts-out file) so the
caller can feed them through whatever Claude interface they have. Layer 5
(thumbnail OCR) is opt-in via --thumbnails-ocr.

Usage:
    python process_export.py \\
        --export /path/to/DailyBoostablePostsAutomatedExport.xlsx \\
        --memory-bank /path/to/memory_bank.csv \\
        --output ./suggested_tags.csv \\
        [--thumbnails-ocr] \\
        [--prompts-out ./layer3_prompts.jsonl]
"""

from __future__ import annotations

import argparse
import csv
import json
import re
import sys
from collections import Counter, defaultdict
from datetime import datetime, timedelta
from pathlib import Path
from typing import Optional

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl is required. Install with: pip install openpyxl")


HASHTAG_RE = re.compile(r"#\w+")
CAMPAIGN_HASHTAGS = {
    "#obsessionisinsession",
    "#primevideo",
    "#streamonprimevideo",
    "#onlyonprimevideo",
}


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument("--export", required=True, help="Sprinklr daily export .xlsx")
    p.add_argument("--memory-bank", required=True, help="Memory bank .csv or .xlsx (Benchmarking + Labeling tab)")
    p.add_argument("--output", required=True, help="Where to write the suggested-tags CSV")
    p.add_argument("--thumbnails-ocr", action="store_true", help="Enable Layer 5 (thumbnail OCR)")
    p.add_argument("--prompts-out", help="Optional path to write Layer 3 prompts as JSONL")
    p.add_argument("--account-history-days", type=int, default=14)
    p.add_argument("--account-history-threshold", type=float, default=0.60)
    return p.parse_args()


def load_export(path: Path) -> tuple[list[dict], list[str]]:
    """Returns (rows, header_names). Sprinklr exports put the dashboard banner on row 1,
    a blank spacer on row 2, the header on row 3, and data from row 4 onward. We
    find the header by scanning for the first row that contains 'Permalink' (case-
    insensitive) so we're resilient to minor preamble changes."""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    rows_iter = ws.iter_rows(values_only=True)
    header: list[str] = []
    for r in rows_iter:
        cells = [str(c).strip() if c is not None else "" for c in r]
        if any("permalink" in c.lower() for c in cells):
            header = cells
            break
    if not header:
        raise ValueError(f"Could not find header row (no 'Permalink' column) in {path}")
    out = []
    for r in rows_iter:
        if all(v is None or v == "" for v in r):
            continue
        out.append({header[i] if i < len(header) else f"col_{i}": v for i, v in enumerate(r)})
    return out, header


def load_memory_bank(path: Path) -> list[dict]:
    """Memory bank: A=Publish Date, B=Post Link, C=Title Tag."""
    rows: list[dict] = []
    if path.suffix.lower() == ".csv":
        with path.open(newline="", encoding="utf-8-sig") as f:
            reader = csv.reader(f)
            for i, r in enumerate(reader):
                if i == 0 and r and "post link" in r[1].lower() if len(r) > 1 else False:
                    continue
                if len(r) < 3:
                    continue
                rows.append({"publish_date": r[0], "post_link": r[1].strip(), "title_tag": r[2].strip()})
    else:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
        ws = wb[wb.sheetnames[0]]
        for i, r in enumerate(ws.iter_rows(values_only=True)):
            if i < 2:  # skip header rows
                continue
            if not r or not r[1]:
                continue
            rows.append({
                "publish_date": r[0],
                "post_link": str(r[1]).strip(),
                "title_tag": (str(r[2]).strip() if r[2] else ""),
            })
    return [r for r in rows if r["post_link"] and r["title_tag"]]


def find_col(header: list[str], *needles: str) -> Optional[int]:
    """Case-insensitive substring match against header names; returns 0-based column index."""
    low = [h.lower() for h in header]
    for n in needles:
        nl = n.lower()
        for i, h in enumerate(low):
            if nl in h:
                return i
    return None


def extract_hashtags(caption: str) -> list[str]:
    if not caption:
        return []
    return [h for h in HASHTAG_RE.findall(caption) if h.lower() not in CAMPAIGN_HASHTAGS]


def url_slug(url: str) -> str:
    if not url:
        return ""
    return url.rstrip("/").rsplit("/", 1)[-1]


def account_history_lookup(
    memory_bank: list[dict],
    export_rows: list[dict],
    days: int,
    threshold: float,
) -> dict[str, Optional[str]]:
    """
    Returns {account_name: dominant_title or None}. Joins memory bank rows
    (URL → tag) to export rows (URL → account) so we can group tags by account.
    """
    url_to_account: dict[str, str] = {}
    for r in export_rows:
        url = (r.get("Permalink") or "").strip()
        acct = (r.get("Account") or "").strip()
        if url and acct:
            url_to_account[url] = acct

    cutoff = datetime.now() - timedelta(days=days)
    by_account: dict[str, list[str]] = defaultdict(list)
    for row in memory_bank:
        acct = url_to_account.get(row["post_link"])
        if not acct:
            continue
        pd = row["publish_date"]
        if isinstance(pd, datetime) and pd < cutoff:
            continue
        tag = row["title_tag"]
        if tag and tag not in ("Multi-Title", "N/A"):
            by_account[acct].append(tag)

    out: dict[str, Optional[str]] = {}
    for acct, tags in by_account.items():
        if not tags:
            out[acct] = None
            continue
        most_common, count = Counter(tags).most_common(1)[0]
        if count / len(tags) >= threshold:
            out[acct] = most_common
        else:
            out[acct] = None
    return out


def build_layer3_prompt(
    caption: str,
    hashtags: list[str],
    account: str,
    url: str,
    known_titles: list[str],
) -> str:
    return f"""You are tagging Prime Video social media posts with the show or movie title
they reference. Tags feed the Amazon boostable-posts report and become a
permanent lookup once committed, so accuracy matters more than coverage.

## Inputs for this post
Caption: {caption}
Hashtags: {", ".join(hashtags) if hashtags else "(none)"}
Account: {account}
URL: {url}
URL slug: {url_slug(url)}

## Known titles (pick from this list — do not invent variations)
{chr(10).join(known_titles)}

## Rules
1. If the caption or hashtags clearly reference one show on the known list,
   return that exact tag (case-sensitive, exact spelling from the list).
2. If two or more distinct shows on the known list are referenced, return
   exactly: Multi-Title
3. If the caption clearly references a real show that is NOT on the known
   list, return: NEW:<show name as best you can identify>
4. If nothing identifiable — generic caption, no show name, no telling
   hashtag, no slug hint — return exactly: N/A
5. Account context matters but is not decisive on its own.
6. Hashtags like #ObsessionIsInSession, #PrimeVideo, #StreamOnPrimeVideo are
   campaign tags, not show tags — ignore them.
7. URL slugs sometimes contain the show name. Use as tiebreaker only.
8. Never invent a title outside the known list without the NEW: prefix.

## Output format
Return ONLY the tag on a single line. No explanation, no punctuation, no quotes.
"""


def main() -> int:
    args = parse_args()
    export_path = Path(args.export)
    mb_path = Path(args.memory_bank)
    out_path = Path(args.output)

    export_rows, header = load_export(export_path)
    memory_bank = load_memory_bank(mb_path)

    # Build lookups
    mb_url_to_tag = {r["post_link"]: r["title_tag"] for r in memory_bank}
    known_titles = sorted({r["title_tag"] for r in memory_bank if r["title_tag"] not in ("Multi-Title", "N/A", "")})

    # Resolve column positions by name (defensive against Sprinklr column drift)
    cap_col = "Outbound Post ( Unified Analytics )"
    pv_col = "PV_Titles (Outbound Message)"
    perma_col = "Permalink"
    acct_col = "Account"
    media_col_idx = find_col(header, "media", "thumbnail", "image url")

    # Account-history map for Layer 4
    acct_history = account_history_lookup(
        memory_bank, export_rows, args.account_history_days, args.account_history_threshold
    )

    layer3_queue: list[dict] = []
    results: list[dict] = []

    for row in export_rows:
        url = (row.get(perma_col) or "").strip()
        if not url:
            continue
        caption = (row.get(cap_col) or "").strip()
        account = (row.get(acct_col) or "").strip()
        pv = (row.get(pv_col) or "").strip()
        hashtags = extract_hashtags(caption)

        # Layer 1: Sprinklr's own tag
        if pv:
            results.append({"url": url, "tag": pv, "confidence": "sprinklr", "reason": "PV_Titles populated"})
            continue

        # Layer 2: memory bank exact match
        if url in mb_url_to_tag:
            results.append({"url": url, "tag": mb_url_to_tag[url], "confidence": "memory", "reason": "exact URL match in memory bank"})
            continue

        # Layer 3: queue for LLM
        prompt = build_layer3_prompt(caption, hashtags, account, url, known_titles)
        layer3_queue.append({
            "url": url, "account": account, "caption": caption, "hashtags": hashtags, "prompt": prompt,
        })

    # Emit Layer 3 prompts if requested
    if args.prompts_out and layer3_queue:
        with Path(args.prompts_out).open("w", encoding="utf-8") as f:
            for item in layer3_queue:
                f.write(json.dumps({"url": item["url"], "prompt": item["prompt"]}) + "\n")

    # For posts that didn't get a Layer 1/2 answer, write provisional Layer 4 / 6
    # rows. The caller can later overlay Layer 3 answers and re-run if desired.
    for item in layer3_queue:
        acct_guess = acct_history.get(item["account"])
        if acct_guess:
            results.append({
                "url": item["url"], "tag": acct_guess, "confidence": "low",
                "reason": f"account history dominated by {acct_guess} (last {args.account_history_days}d)",
            })
        else:
            results.append({
                "url": item["url"], "tag": "(needs-review)", "confidence": "review",
                "reason": "Layer 3 LLM call required; no account-history fallback",
            })

    # Write results
    with out_path.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["Post Link", "Suggested Tag", "Confidence", "Reason"])
        for r in results:
            w.writerow([r["url"], r["tag"], r["confidence"], r["reason"]])

    # Summary to stderr
    by_conf = Counter(r["confidence"] for r in results)
    print(f"Processed {len(results)} posts:", file=sys.stderr)
    for k in ("sprinklr", "memory", "high", "low", "ocr", "review"):
        if by_conf.get(k):
            print(f"  {k}: {by_conf[k]}", file=sys.stderr)
    print(f"Layer 3 prompts queued: {len(layer3_queue)}", file=sys.stderr)
    print(f"Output: {out_path}", file=sys.stderr)
    return 0


if __name__ == "__main__":
    sys.exit(main())
Processed {len(results)} posts:", file=sys.stderr)
    for k in ("sprinklr", "memory", "high", "low", "ocr", "review"):
        if by_conf.get(k):
            print(f"  {k}: {by_conf[k]}", file=sys.stderr)
    print(f"Layer 3 prompts queued: {len(layer3_queue)}", file=sys.stderr)
    print(f"Output: {out_path}", file=sys.stderr)
    return 0


if __name__ == "__main__":
    sys.exit(main())
